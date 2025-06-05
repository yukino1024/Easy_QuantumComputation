from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill,Alignment

class wri_excel:
    def __init__(self):
        self.header =  ["", "RASSCF", "ΔE(ras)", "CASPT2", "ΔE(pt2)", "wavelength", "Osc.", "Occ/unOcc", "state", "D.M."]
        self.data = {}
        self.wb = Workbook()
        self.ws = self.wb.active
        self._wri_header()
        self._set_width()


    def _wri_header(self):
        for col_num, col_name in enumerate(self.header, start=1):
            self.ws.cell(row=1, column=col_num, value=col_name)
        for cell in self.ws['1']:
            cell.font = Font(name="DengXian", bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(fill_type='solid', fgColor='FFE69AE6')

    def wri_data(self):
        for raw_num in range(1,len(self.data)+1):
            for col_num,col_name in enumerate(self.data[str(raw_num)], start=1):
                self.ws.cell(row=raw_num+1, column=col_num, value=col_name)
            self._set_higth(raw_num+1, 15)

    def save_excel(self, filename):
        self._set_number()
        self.wb.save(filename)

    def add_data(self, i,ras_data,pt2_data,osc_data,occ_data):
        if osc_data == '':
            if pt2_data == 'Empty':
                self.data[str(i)] = [str(i),self._to_float(ras_data),'=(B'+str(i+1)+'-B2)*627.5',pt2_data,'Empty','Empty',osc_data,occ_data]
            else:
                self.data[str(i)] = [str(i),self._to_float(ras_data),'=(B'+str(i+1)+'-B2)*627.5',self._to_float(pt2_data),'=(D'+str(i+1)+'-D2)*627.5','=28591/'+'E'+str(i+1),osc_data,occ_data]
        else:
            if pt2_data == 'Empty':
                self.data[str(i)] = [str(i),self._to_float(ras_data),'=(B'+str(i+1)+'-B2)*627.5',pt2_data,'Empty','Empty',self._to_float(osc_data),occ_data]
            else:
                self.data[str(i)] = [str(i),self._to_float(ras_data),'=(B'+str(i+1)+'-B2)*627.5',self._to_float(pt2_data),'=(D'+str(i+1)+'-D2)*627.5','=28591/'+'E'+str(i+1),self._to_float(osc_data),occ_data]


    #设置行高列宽
    def _set_width(self):
        self.ws.column_dimensions['A'].width = 7
        self.ws.column_dimensions['B'].width = 14
        self.ws.column_dimensions['C'].width = 8
        self.ws.column_dimensions['D'].width = 14
        self.ws.column_dimensions['E'].width = 8
        self.ws.column_dimensions['F'].width = 12
        self.ws.column_dimensions['G'].width = 16
        self.ws.column_dimensions['H'].width = 16
        self.ws.column_dimensions['I'].width = 10
        self.ws.column_dimensions['J'].width = 18


    def _set_higth(self,row_num,height):
        self.ws.row_dimensions[row_num].height = height
        for cell in self.ws[row_num]:
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name="DengXian", bold=True)


    #设置为数字格式
    def _set_number(self):
        for cell in self.ws['B']:
            cell.number_format = '0.0000'
        for cell in self.ws['C']:
            cell.number_format = '0.00'
        for cell in self.ws['D']:
            cell.number_format = '0.0000'
        for cell in self.ws['E']:
            cell.number_format = '0.00'
        for cell in self.ws['F']:
            cell.number_format = '0'
        for cell in self.ws['G']:
            cell.number_format = '0.00E+00'

    #转float
    def _to_float(self, value):
        try:
            return float(value)
        except:
            return value  # 如果转换失败，返回原始值




